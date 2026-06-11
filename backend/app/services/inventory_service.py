"""Lógica de inventario: serialización JSON, historial y Excel (Fase 7)."""
import io
import json
import unicodedata
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from .. import models


def _load_json(value: str | None, default):
    if not value:
        return default
    try:
        return json.loads(value)
    except (ValueError, TypeError):
        return default


def category_to_out(category: models.InventoryCategory) -> dict:
    return {
        "id": category.id,
        "name": category.name,
        "description": category.description,
        "icon": category.icon,
        "fields_schema": _load_json(category.fields_schema, []),
        "created_at": category.created_at,
        "item_count": len(category.items),
    }


def item_to_out(item: models.InventoryItem) -> dict:
    return {
        "id": item.id,
        "category_id": item.category_id,
        "category_name": item.category.name if item.category else None,
        "name": item.name,
        "description": item.description,
        "serial_number": item.serial_number,
        "brand": item.brand,
        "model": item.model,
        "status": item.status,
        "location": item.location,
        "assigned_to": item.assigned_to,
        "device_id": item.device_id,
        "purchase_date": item.purchase_date,
        "warranty_until": item.warranty_until,
        "notes": item.notes,
        "custom_fields": _load_json(item.custom_fields, {}),
        "photo_url": item.photo_url,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def log_history(
    db: Session,
    item_id: int,
    action: str,
    details: str | None,
    changed_by: str | None,
) -> None:
    db.add(
        models.InventoryHistory(
            item_id=item_id, action=action, details=details, changed_by=changed_by
        )
    )


# ---------------------------------------------------------------------------
# Exportar / importar Excel
# ---------------------------------------------------------------------------
def build_workbook(items: list[models.InventoryItem], category_name: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    
    # Determinar columnas y título de hoja basados en la categoría
    if category_name == "Correos":
        cols = [
            ("id", "ID"),
            ("assigned_to", "Nombre"),
            ("name", "Correo"),
            ("model", "Licencia"),
            ("status", "Estado"),
        ]
        ws.title = "Correos"
    elif category_name == "Licencias":
        cols = [
            ("id", "ID"),
            ("name", "Nombre / Activo"),
            ("brand", "Editor"),
            ("model", "Modelo"),
            ("serial_number", "Clave (Product Key)"),
            ("assigned_to", "Asignado a"),
            ("status", "Estado"),
        ]
        ws.title = "Licencias"
    elif category_name == "Software":
        cols = [
            ("id", "ID"),
            ("name", "Nombre / Activo"),
            ("brand", "Desarrollador"),
            ("model", "Versión"),
            ("serial_number", "Nro Serie"),
            ("assigned_to", "Asignado a"),
            ("status", "Estado"),
        ]
        ws.title = "Software"
    elif category_name:
        cols = [
            ("id", "ID"),
            ("name", "Nombre / Activo"),
            ("brand", "Marca"),
            ("model", "Modelo"),
            ("serial_number", "Nro Serie"),
            ("assigned_to", "Asignado a"),
            ("status", "Estado"),
            ("location", "Ubicación"),
            ("purchase_date", "Fecha compra"),
            ("warranty_until", "Garantía hasta"),
            ("notes", "Notas"),
        ]
        import re
        ws.title = re.sub(r'[^a-zA-Z0-9_\-\s]', '_', category_name)[:30]
    else:
        # Exportación global (todas las categorías)
        cols = [
            ("id", "ID"),
            ("category_name", "Categoría"),
            ("name", "Nombre"),
            ("serial_number", "Nro Serie"),
            ("brand", "Marca"),
            ("model", "Modelo"),
            ("status", "Estado"),
            ("location", "Ubicación"),
            ("assigned_to", "Asignado a"),
            ("purchase_date", "Fecha compra"),
            ("warranty_until", "Garantía hasta"),
            ("notes", "Notas"),
        ]
        ws.title = "Inventario General"

    status_labels = {
        "active": "Activo",
        "in_repair": "En mantenimiento",
        "retired": "De baja",
        "lost": "Perdido",
    }

    ws.append([label for _, label in cols])
    for item in items:
        row = item_to_out(item)
        values = []
        for key, _ in cols:
            val = row.get(key)
            if key == "status" and val in status_labels:
                val = status_labels[val]
            elif isinstance(val, (date, datetime)):
                val = val.isoformat()
            values.append(val)
        ws.append(values)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _normalize(text: str) -> str:
    """Normaliza un string quitando acentos y convirtiendo a minúsculas.

    Permite que 'Ubicación' matchee con 'ubicacion', 'Nro Serie' con 'nro serie', etc.
    """
    text = text.lower().strip()
    # Descompone caracteres unicode (é → e + acento) y quita los diacríticos.
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _str_or_none(value) -> str | None:
    """Convierte un valor a string limpio, o None si está vacío."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def import_workbook(
    db: Session, content: bytes, default_category_id: int, changed_by: str | None
) -> dict:
    """Importa items desde un Excel (multi-hoja). Hace upsert por nro de serie o nombre."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    
    created = updated = 0
    
    # Recorrer todas las hojas del archivo Excel
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Buscar fila de encabezado que contenga columnas conocidas
        header_index = -1
        header = []
        known_headers = {
            "nombre", "name", "equipo", "activo", "item", "nro serie", "nro. serie", 
            "numero serie", "serial", "serie", "marca", "brand", "modelo", "model",
            "estado", "status", "ubicacion", "ubicación", "location", "correo", "email", "mail",
            "licencia", "license", "licence"
        }

        for idx, row in enumerate(rows):
            if not row:
                continue
            normalized_row = [_normalize(str(cell).strip()) if cell is not None else "" for cell in row]
            if any(h in known_headers for h in normalized_row):
                header_index = idx
                header = normalized_row
                break
        
        if header_index == -1:
            # Fallback: primera fila con algún contenido
            for idx, row in enumerate(rows):
                if row and any(c is not None for c in row):
                    header_index = idx
                    header = [_normalize(str(cell).strip()) if cell is not None else "" for cell in row]
                    break
            else:
                continue

        def col(row, *names):
            for n in names:
                normalized = _normalize(n)
                if normalized in header:
                    val = row[header.index(normalized)]
                    return val
            return None

        for row in rows[header_index + 1:]:
            if not row or all(c is None for c in row):
                continue
            
            # Mapeo inteligente de correo
            correo_val = col(row, "correo", "email", "mail", "correo electronico", "correo electrónico")
            nombre_val = col(row, "nombre", "name", "equipo", "activo", "item", "descripcion", "descripción")
            
            if correo_val:
                name = str(correo_val).strip()
                assigned_to = str(nombre_val).strip() if nombre_val else None
            else:
                name = nombre_val
                if not name:
                    # Fallback inteligente: primer valor no vacío de la fila
                    non_empty_cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if non_empty_cells:
                        name = non_empty_cells[0]
                    else:
                        continue
                name = str(name).strip()
                assigned_to = _str_or_none(col(row, "asignado a", "asignado", "assigned_to", "usuario", "responsable"))
            
            serial = col(row, "nro serie", "nro. serie", "numero serie", "número serie",
                          "serial_number", "serial", "serie", "s/n")

            existing = None
            if serial:
                serial_str = str(serial).strip()
                existing = (
                    db.query(models.InventoryItem)
                    .filter(models.InventoryItem.serial_number == serial_str)
                    .first()
                )
            if not existing:
                existing = (
                    db.query(models.InventoryItem)
                    .filter(models.InventoryItem.name == name)
                    .first()
                )

            # Para modelo, usar el nombre de la hoja (licencia) como fallback si no viene explícito en columnas
            model_val = _str_or_none(col(row, "modelo", "model", "licencia", "license", "licence"))
            if not model_val:
                model_val = sheet_name.strip()

            fields = dict(
                name=name,
                serial_number=str(serial).strip() if serial else None,
                brand=_str_or_none(col(row, "marca", "brand", "fabricante", "manufacturer")),
                model=model_val,
                status=_str_or_none(col(row, "estado", "status")) or "active",
                location=_str_or_none(col(row, "ubicación", "ubicacion", "location", "sede", "oficina")),
                assigned_to=assigned_to,
                purchase_date=_parse_date(col(row, "fecha compra", "fecha de compra", "purchase_date")),
                warranty_until=_parse_date(col(row, "garantía hasta", "garantia hasta",
                                               "warranty_until", "fin garantía", "fin garantia")),
                notes=_str_or_none(col(row, "notas", "notes", "observaciones", "comentarios", "descripción", "descripcion")),
            )

            # Si el estado viene en español (ej. "Activo"), normalizarlo a minúsculas
            if fields["status"]:
                status_lower = fields["status"].lower().strip()
                if status_lower in ("activo", "active"):
                    fields["status"] = "active"
                elif status_lower in ("reparacion", "reparación", "in_repair"):
                    fields["status"] = "in_repair"
                elif status_lower in ("retirado", "retired"):
                    fields["status"] = "retired"
                elif status_lower in ("perdido", "lost"):
                    fields["status"] = "lost"

            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                item = models.InventoryItem(category_id=default_category_id, **fields)
                db.add(item)
                db.flush()
                log_history(db, item.id, "created", "Importado desde Excel", changed_by)
                created += 1

    db.commit()
    return {"created": created, "updated": updated}
